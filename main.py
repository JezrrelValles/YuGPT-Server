from collections import Counter
import os
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Body
from openai import AsyncOpenAI, OpenAIError
from openai.types.beta.threads.run import RequiredAction, LastError
from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from io import BytesIO
from typing import Any, Dict, List, Optional
import pandas as pd
import asyncio
from bank_processors.bank_processor_factory import BankProcessorFactory
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from mistralai import Mistral
import datetime
import math
import calendar
from datetime import datetime
import traceback

app = FastAPI()
load_dotenv()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

client = AsyncOpenAI(
    api_key=os.getenv("API_KEY"),
)

mistral_api_key = os.getenv("MISTRAL_API")
mistral_client = Mistral(api_key=mistral_api_key)

MESES = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}

MESES_REVERSO = {v: k for k, v in MESES.items()}

BANKS_TO_ASSISTANT_ID = {
    "BANAMEX": "asst_1M8w2HKrJqJdjkPji7eF7JJK",
    "BANBAJIO": "asst_YRhqGDSFvH8K5siImwzl32sx",
    "BANORTE": "asst_dHWzuDg0D0rxyXGay3hC9uEM",
    "BANREGIO": "asst_h9ScdBXkSI3V5yV7w1BTg7f5",
    "BBVA": "asst_Bmuha99CJW515evmB4vljYH7",
    "BX+": "asst_mLkxpzOSBF5YadZiiZhC4y3U",
    "CHASE": "asst_68agxxAZQii7q0vS5oKS4L1s",
    "HSBC": "asst_9Mzw7u5Z684PnipVujq1kNxa",
    "SANTANDER": "asst_HOQwBr6KtWHAfv2JbrNySRN5",
    "SCOTIABANK": "asst_6lN2cWBcCQ8RuiOqWWA7clgu",
}

run_finished_states = ["completed", "failed", "cancelled", "expired", "requires_action"]


class RunStatus(BaseModel):
    run_id: str
    thread_id: str
    status: str
    required_action: Optional[RequiredAction]
    last_error: Optional[LastError]


class ThreadMessage(BaseModel):
    content: str
    role: str
    hidden: bool
    id: str
    created_at: int


class Thread(BaseModel):
    messages: List[ThreadMessage]


class CreateMessage(BaseModel):
    content: str


class TransaccionGroup(BaseModel):
    total: float
    transacciones: List[List[float]]


class Saldos(BaseModel):
    saldo_contabilidad: float
    saldo_libros: float
    saldo_estado_cuenta: float
    saldo_bancos: float
    diferencia: float


class PreviousConciliationResult(BaseModel):
    empresa: str
    mes: str
    cuenta: str
    saldos: Saldos
    depositos: TransaccionGroup
    retiros: TransaccionGroup
    depositos_en_transito: TransaccionGroup
    cheques_en_transito: TransaccionGroup


class Movimiento(BaseModel):
    fecha: str
    tipo: str
    monto: float
    saldo: float


class AuxResult(BaseModel):
    saldo_inicial: float
    datos: List[Movimiento]
    total_cargos: float
    total_abonos: float
    total_saldo: float


class ConciliationRequest(BaseModel):
    previousConciliationResult: PreviousConciliationResult
    auxResult: AuxResult
    assistantResult: List[Movimiento]


def obtener_descripcion(prev_fecha_str):
    mes_str, anio_str = prev_fecha_str.split()
    mes = MESES.get(mes_str.upper())
    anio = int(anio_str)

    if mes == 12:
        mes_siguiente = 1
        anio += 1
    else:
        mes_siguiente = mes + 1

    ultimo_dia = calendar.monthrange(anio, mes_siguiente)[1]

    nombre_mes = MESES_REVERSO[mes_siguiente]

    return f"SALDO EN CONTABILIDAD AL {ultimo_dia} DE {nombre_mes} DEL {anio}"


def extraer_movimientos_conciliacion(prev) -> List[Movimiento]:
    movimientos = []

    for lista in prev.depositos.transacciones:
        for monto in lista:
            movimientos.append(Movimiento(monto=monto, tipo="deposito"))

    for lista in prev.retiros.transacciones:
        for monto in lista:
            movimientos.append(Movimiento(monto=monto, tipo="retiro"))

    for lista in prev.depositos_en_transito.transacciones:
        for monto in lista:
            movimientos.append(Movimiento(monto=monto, tipo="deposito"))

    for lista in prev.cheques_en_transito.transacciones:
        for monto in lista:
            movimientos.append(Movimiento(monto=monto, tipo="retiro"))

    return movimientos


def obtener_siguiente_fecha(prev_fecha_str):
    mes_str, anio_str = prev_fecha_str.split()
    mes = MESES.get(mes_str.upper())
    anio = int(anio_str)

    if mes == 12:
        mes_siguiente = 1
        anio_siguiente = anio + 1
    else:
        mes_siguiente = mes + 1
        anio_siguiente = anio

    nombre_mes = MESES_REVERSO[mes_siguiente]

    return f"{nombre_mes} {anio_siguiente}"


def comparar_transacciones(
    assistant: List[Movimiento], aux: AuxResult, prev_movimientos: List[Movimiento]
) -> Dict[str, Any]:
    assistant_counts = Counter((item.monto, item.tipo) for item in assistant)
    aux_counts = Counter((item.monto, item.tipo) for item in aux.datos)
    prev_counts = Counter((item.monto, item.tipo) for item in prev_movimientos)

    match_transactions = {}
    assistant_discrepancies = {}
    aux_discrepancies = {}
    prev_discrepancies = {}

    for transaction, count in assistant_counts.items():
        if transaction in aux_counts:
            match_count = min(count, aux_counts[transaction])

            if match_count > 0:
                match_transactions[transaction] = match_count

            assistant_discrepancies[transaction] = count - match_count
            aux_counts[transaction] -= match_count
        else:
            assistant_discrepancies[transaction] = count

    for transaction, count in aux_counts.items():
        if count > 0:
            aux_discrepancies[transaction] = count

    for transaction, count in prev_counts.items():
        if transaction in assistant_discrepancies:
            match_count = min(count, assistant_discrepancies[transaction])
            if match_count > 0:
                match_transactions[transaction] = (
                    match_transactions.get(transaction, 0) + match_count
                )
                assistant_discrepancies[transaction] -= match_count
                prev_counts[transaction] -= match_count

    for transaction, count in prev_counts.items():
        if count > 0:
            prev_discrepancies[transaction] = count

    match_transactions = {k: v for k, v in match_transactions.items() if v > 0}
    assistant_discrepancies = {
        k: v for k, v in assistant_discrepancies.items() if v > 0
    }
    aux_discrepancies = {k: v for k, v in aux_discrepancies.items() if v > 0}
    prev_discrepancies = {k: v for k, v in prev_discrepancies.items() if v > 0}

    return {
        "matches": dict(sorted(match_transactions.items())),
        "assistant_discrepancies": dict(sorted(assistant_discrepancies.items())),
        "aux_discrepancies": dict(sorted(aux_discrepancies.items())),
        "prev_discrepancies": dict(sorted(prev_discrepancies.items())),
    }


def get_assistant_id(bank: str) -> str:
    if bank not in BANKS_TO_ASSISTANT_ID:
        raise HTTPException(status_code=400, detail="Banco no válido")
    return BANKS_TO_ASSISTANT_ID[bank]


async def convert_pdf_to_text(pdf_path, bank):
    with open(pdf_path, "rb") as pdf_file:
        processor = BankProcessorFactory.get_processor(bank, pdf_file)
        process_data = processor.process()
        process_data = "\n".join(
            [" ".join(map(str, row)).strip() for row in process_data]
        )

    return process_data


async def convert_scanned_pdf_to_text(pdf_path):
    try:
        with open(pdf_path, "rb") as file:
            uploaded_pdf = mistral_client.files.upload(
                file={
                    "file_name": pdf_path,
                    "content": file,
                },
                purpose="ocr",
            )

        signed_url = mistral_client.files.get_signed_url(file_id=uploaded_pdf.id)
        print(signed_url)
        ocr_response_pages = mistral_client.ocr.process(
            model="mistral-ocr-latest",
            document={"type": "document_url", "document_url": signed_url.url},
        ).pages

        text = ""
        for page in ocr_response_pages:
            text += page.markdown + "\n"

        print(f"Texto : {text}")

        await mistral_client.files.delete_async(file_id=uploaded_pdf.id)

        # Create a .txt with the ocr text extracted
        with open(
            datetime.now().strftime("%Y%m%d%H%M%S") + ".txt",
            "w",
            encoding="utf-8",
        ) as file:
            file.write(text)

        return text
    except Exception as e:
        raise Exception(f"Error en convert_scanned_pdf_to_text: {str(e)}")


async def extract_text_from_aux(file: UploadFile):
    try:
        content = await file.read()  # Read file as binary
        excel_data = BytesIO(content)  # Convert to BytesIO for pandas

        # Read the file starting from row 7 (row index 6 in pandas)
        df = pd.read_excel(excel_data, skiprows=6, usecols="A,E:H")

        # Rename columns manually
        df.columns = ["Fecha", "Referencia", "Cargos", "Abonos", "Saldo"]

        saldo_inicial = df.iloc[1]["Saldo"]

        df_data = df.iloc[4:].reset_index(drop=True)

        df_data = df[
            df["Fecha"].replace([None, "", " ", 0], pd.NA).notna()
        ].reset_index(drop=True)

        df_data["Cargos"] = pd.to_numeric(df_data["Cargos"], errors="coerce").fillna(0)
        df_data["Abonos"] = pd.to_numeric(df_data["Abonos"], errors="coerce").fillna(0)
        df_data["Saldo"] = pd.to_numeric(df_data["Saldo"], errors="coerce").fillna(0)

        total_cargos = round(df_data["Cargos"].sum(), 2)
        total_abonos = round(df_data["Abonos"].sum(), 2)
        total_saldo = round(saldo_inicial + (total_cargos - total_abonos), 2)

        # Convert data to JSON format
        extracted_text = {
            "saldo_inicial": saldo_inicial,
            "datos": [
                {
                    "fecha": row["Fecha"],
                    "tipo": (
                        "saldo inicial"
                        if row["Cargos"] == 0 and row["Abonos"] == 0
                        else "retiro" if row["Abonos"] > 0 else "deposito"
                    ),
                    "monto": row["Abonos"] if row["Abonos"] > 0 else row["Cargos"],
                    "saldo": row["Saldo"],
                }
                for _, row in df_data.iterrows()
            ],
            "total_cargos": total_cargos,
            "total_abonos": total_abonos,
            "total_saldo": total_saldo,
        }

        print(extracted_text)
        return extracted_text
    except Exception as e:
        return {"error": f"Error processing the Excel file: {str(e)}"}


def is_nan(value):
    try:
        return math.isnan(value)
    except:
        return False


async def extract_text_from_previous(file: UploadFile):
    try:
        # Leer el archivo
        content = await file.read()
        previous_data = BytesIO(content)
        df = pd.read_excel(previous_data, header=None)

        # Convertir DataFrame a lista de listas
        data = df.values.tolist()

        # Extraer información clave
        empresa = data[0][2] if len(data[0]) > 1 else None
        mes = data[1][2] if len(data[1]) > 1 else None
        cuenta = data[2][2] if len(data[2]) > 1 else None

        saldo_contabilidad = (
            data[5][-1] if isinstance(data[5][-1], (int, float)) else None
        )
        saldo_estado_cuenta = (
            data[27][-1] if isinstance(data[27][-1], (int, float)) else None
        )
        saldo_bancos = data[46][-1] if isinstance(data[46][-1], (int, float)) else None
        saldo_libros = data[24][-1] if isinstance(data[24][-1], (int, float)) else None

        depositos = data[8][-1] if isinstance(data[8][-1], (int, float)) else None
        retiros = data[16][-1] if isinstance(data[16][-1], (int, float)) else None
        depositos_en_transito = (
            data[30][-1] if isinstance(data[30][-1], (int, float)) else None
        )
        cheques_en_transito = (
            data[38][-1] if isinstance(data[38][-1], (int, float)) else None
        )
        diferencia = data[47][-1] if isinstance(data[47][-1], (int, float)) else None

        transacciones_depositos = [
            [item for item in data[i][0:7:2] if not is_nan(item)] for i in range(9, 15)
        ]
        transacciones_retiros = [
            [item for item in data[i][0:7:2] if not is_nan(item)] for i in range(17, 23)
        ]
        transacciones_depositos_en_transito = [
            [item for item in data[i][0:7:2] if not is_nan(item)] for i in range(31, 37)
        ]
        transacciones_cheques_en_transito = [
            [item for item in data[i][0:7:2] if not is_nan(item)] for i in range(39, 45)
        ]

        # Construcción del diccionario final
        extracted_text = {
            "empresa": empresa,
            "mes": mes,
            "cuenta": cuenta,
            "saldos": {
                "saldo_contabilidad": saldo_contabilidad,
                "saldo_libros": saldo_libros,
                "saldo_estado_cuenta": saldo_estado_cuenta,
                "saldo_bancos": saldo_bancos,
                "diferencia": diferencia,
            },
            "depositos": {
                "total": depositos,
                "transacciones": transacciones_depositos,
            },
            "retiros": {
                "total": retiros,
                "transacciones": transacciones_retiros,
            },
            "depositos_en_transito": {
                "total": depositos_en_transito,
                "transacciones": transacciones_depositos_en_transito,
            },
            "cheques_en_transito": {
                "total": cheques_en_transito,
                "transacciones": transacciones_cheques_en_transito,
            },
        }
        print(extracted_text)
        return extracted_text
    except Exception as e:
        return {"error": f"Error processing the Excel file: {str(e)}"}


async def wait_on_run(thread_id: str, run_id: str, polling_interval: int = 3):
    """
    Waits for an OpenAI run to complete.

    Args:
        thread_id (str): The ID of the thread.
        run_id (str): The ID of the run.
        polling_interval (int): Time in seconds to wait between API checks.

    Returns:
        RunStatus: The final status of the run.
    """
    while True:
        run = await client.beta.threads.runs.retrieve(
            thread_id=thread_id, run_id=run_id
        )

        if run.status in run_finished_states:
            return RunStatus(
                run_id=run.id,
                thread_id=thread_id,
                status=run.status,
                required_action=run.required_action,
                last_error=run.last_error,
            )

        await asyncio.sleep(polling_interval)


@app.post("/extract_account/")
async def extract_account(file: UploadFile = File(...), bank: str = Form(...)):
    assistant = get_assistant_id(bank)
    file_location = f"temp_{file.filename}"

    try:
        with open(file_location, "wb") as f:
            f.write(await file.read())

        try:
            text = await convert_pdf_to_text(file_location, bank.lower())
        except Exception as e:
            print(f"Error en convert_pdf_to_text: {e}")
            text = None

        if not text:
            try:
                text = await convert_scanned_pdf_to_text(file_location)
            except Exception as e:
                print(f"Error en convert_scanned_pdf_to_text: {e}")
                return JSONResponse(
                    content={"error": f"Error al procesar PDF: {str(e)}"},
                    status_code=500,
                )

        return {
            "assistant": assistant,
            "filename": file.filename,
            "extracted_text": text,
        }
    finally:
        if os.path.exists(file_location):
            os.remove(file_location)


@app.post("/extract_aux/")
async def extract_aux(file: UploadFile = File(...)):
    try:
        text_data = await extract_text_from_aux(file)

        if isinstance(text_data, dict) and "error" in text_data:
            return JSONResponse(content=text_data, status_code=400)

        return {"aux_transactions": text_data}
    except Exception as e:
        return JSONResponse(
            content={"error": f"Error al procesar el archivo Excel: {str(e)}"},
            status_code=500,
        )


@app.post("/extract_previous/")
async def extract_previous(file: UploadFile = File(...)):
    try:
        text_data = await extract_text_from_previous(file)

        if isinstance(text_data, dict) and "error" in text_data:
            return JSONResponse(content=text_data, status_code=400)

        return {"previous_transactions": text_data}
    except Exception as e:
        return JSONResponse(
            content={"error": f"Error al procesar el archivo Excel: {str(e)}"},
            status_code=500,
        )


@app.post("/api/new")
async def post_new(data: dict = Body(...)):
    try:
        extracted_text = data.get("extracted_text")
        assistant = data.get("assistant")

        if not extracted_text or not assistant:
            raise HTTPException(status_code=400, detail="Faltan datos en la solicitud.")

        thread = await client.beta.threads.create()

        message = await client.beta.threads.messages.create(
            thread_id=thread.id, role="user", content=extracted_text
        )

        run = await client.beta.threads.runs.create(
            thread_id=thread.id, assistant_id=assistant
        )

        run_status = await wait_on_run(thread.id, run.id)

        if run_status.status != "completed":
            return JSONResponse(
                content={
                    "error": f"Error al procesar la solicitud: {run_status.status}"
                },
                status_code=500,
            )

        messages = await client.beta.threads.messages.list(thread_id=thread.id)

        assistant_messages = [msg for msg in messages.data if msg.role == "assistant"]

        assistant_response = (
            assistant_messages[-1].content[0].text.value
            if assistant_messages
            else "No hay respuesta del asistente."
        )

        return {"assistant_transactions": assistant_response}
    except OpenAIError as e:
        raise HTTPException(status_code=500, detail=f"Error de OpenAI: {str(e)}")


@app.post("/create_conciliation/")
async def create_conciliation(data: ConciliationRequest):
    yellowHighlighter = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    orangeHighlighter = PatternFill(
        start_color="FFA500", end_color="FFA500", fill_type="solid"
    )

    try:
        prev = data.previousConciliationResult
        aux = data.auxResult
        assistant = data.assistantResult

        prev_movimientos = extraer_movimientos_conciliacion(prev)

        # Prev
        depositos_total = prev.depositos.total if prev.depositos.total > 0 else 0
        retiros_total = prev.retiros.total if prev.retiros.total > 0 else 0
        depositos_en_transito_total = (
            prev.depositos_en_transito.total
            if prev.depositos_en_transito.total > 0
            else 0
        )
        cheques_en_transito_total = (
            prev.cheques_en_transito.total if prev.cheques_en_transito.total > 0 else 0
        )

        fecha_actual = obtener_siguiente_fecha(prev.mes)
        descripcion = obtener_descripcion(prev.mes)

        # Aux
        aux_saldo_final = aux.total_saldo

        # assistant vs aux
        comparetive_assistant_aux_conciliation = comparar_transacciones(
            assistant, aux, prev_movimientos
        )
        print(comparetive_assistant_aux_conciliation)

        saldo_final = 0.0
        penultimo_saldo = 0.0
        encontrado_saldo_final = False

        if assistant:  # Verifica si la lista no está vacía
            encontrado_saldo_final = False
            saldo_final = 0.0
            penultimo_saldo = 0.0

            # Iteramos desde el final para encontrar el último movimiento
            for transaction in reversed(assistant):
                if not encontrado_saldo_final:
                    saldo_final = transaction.saldo
                    encontrado_saldo_final = True
                elif saldo_final == 0.0:
                    penultimo_saldo = transaction.saldo
                    break  # Encontramos el penúltimo, salimos del bucle

            # Si el saldo final es 0, usamos el penúltimo como referencia
            if saldo_final == 0.0 and penultimo_saldo != 0.0:
                saldo_final = penultimo_saldo

        archivo_excel = "format.xlsx"
        wb = load_workbook(archivo_excel)
        hoja = wb.active

        hoja["C1"] = prev.empresa  # empresa
        hoja["C2"] = fecha_actual  # mes
        hoja["C3"] = prev.cuenta  # cuenta
        hoja["B6"] = descripcion

        saldo_contabilidad = aux_saldo_final
        total_depositos = depositos_total
        total_retiros = retiros_total
        saldo_libros = (saldo_contabilidad + total_depositos) - total_retiros
        saldo_estado_cuenta = saldo_final
        total_depositos_transito = depositos_en_transito_total
        total_cheques_transito = cheques_en_transito_total
        saldo_bancos = (
            saldo_estado_cuenta + total_depositos_transito
        ) - total_cheques_transito
        diferencia = saldo_libros - saldo_bancos

        hoja["H6"] = saldo_contabilidad
        # hoja["H9"] = total_depositos
        # hoja["H17"] = total_retiros
        # hoja["H25"] = saldo_libros
        hoja["H28"] = saldo_estado_cuenta
        # hoja["H31"] = total_depositos_transito
        # hoja["H39"] = total_cheques_transito
        # hoja["H47"] = saldo_bancos
        # hoja["H48"] = diferencia

        assistant_discrepancy_rows = {
            "deposito": {
                "col": [
                    "A",
                    "B",
                    "C",
                    "D",
                    "E",
                    "F",
                    "G",
                    "I",
                    "J",
                    "K",
                    "L",
                    "M",
                    "N",
                    "O",
                    "P",
                    "Q",
                    "R",
                    "S",
                    "T",
                    "U",
                    "V",
                    "W",
                    "X",
                    "Y",
                    "Z",
                ],
                "row_start": 10,
                "index": 0,
            },
            "retiro": {
                "col": [
                    "A",
                    "B",
                    "C",
                    "D",
                    "E",
                    "F",
                    "G",
                    "I",
                    "J",
                    "K",
                    "L",
                    "M",
                    "N",
                    "O",
                    "P",
                    "Q",
                    "R",
                    "S",
                    "T",
                    "U",
                    "V",
                    "W",
                    "X",
                    "Y",
                    "Z",
                ],
                "row_start": 18,
                "index": 0,
            },
        }

        for (monto, tipo), count in comparetive_assistant_aux_conciliation[
            "assistant_discrepancies"
        ].items():
            if tipo in assistant_discrepancy_rows:
                info = assistant_discrepancy_rows[tipo]
                for _ in range(count):
                    col_index = info["index"] // 6
                    row_offset = info["index"] % 6

                    if col_index < len(info["col"]):
                        cell = (
                            f"{info['col'][col_index]}{info['row_start'] + row_offset}"
                        )
                        hoja[cell] = monto
                        hoja[cell].fill = yellowHighlighter
                        info["index"] += 1

        # Previos pendientes (naranja, mismas filas que assistant)
        for (monto, tipo), count in comparetive_assistant_aux_conciliation[
            "prev_discrepancies"
        ].items():
            if tipo in assistant_discrepancy_rows:
                info = assistant_discrepancy_rows[tipo]
                for _ in range(count):
                    col_index = info["index"] // 6
                    row_offset = info["index"] % 6

                    if col_index < len(info["col"]):
                        cell = (
                            f"{info['col'][col_index]}{info['row_start'] + row_offset}"
                        )
                        hoja[cell] = monto
                        hoja[cell].fill = orangeHighlighter
                        info["index"] += 1

        aux_discrepancy_rows = {
            "deposito": {
                "col": [
                    "A",
                    "B",
                    "C",
                    "D",
                    "E",
                    "F",
                    "G",
                    "I",
                    "J",
                    "K",
                    "L",
                    "M",
                    "N",
                    "O",
                    "P",
                    "Q",
                    "R",
                    "S",
                    "T",
                    "U",
                    "V",
                    "W",
                    "X",
                    "Y",
                    "Z",
                ],
                "row_start": 32,
                "index": 0,
            },
            "retiro": {
                "col": [
                    "A",
                    "B",
                    "C",
                    "D",
                    "E",
                    "F",
                    "G",
                    "I",
                    "J",
                    "K",
                    "L",
                    "M",
                    "N",
                    "O",
                    "P",
                    "Q",
                    "R",
                    "S",
                    "T",
                    "U",
                    "V",
                    "W",
                    "X",
                    "Y",
                    "Z",
                ],
                "row_start": 40,
                "index": 0,
            },
        }

        for (monto, tipo), count in comparetive_assistant_aux_conciliation[
            "aux_discrepancies"
        ].items():
            if tipo in aux_discrepancy_rows:
                info = aux_discrepancy_rows[tipo]
                for _ in range(count):
                    col_index = info["index"] // 6
                    row_offset = info["index"] % 6

                if col_index < len(info["col"]):
                    cell = f"{info['col'][col_index]}{info['row_start'] + row_offset}"
                    hoja[cell] = monto
                    hoja[cell].fill = yellowHighlighter
                    info["index"] += 1

            # Previos pendientes (naranja, mismas filas que assistant)
        for (monto, tipo), count in comparetive_assistant_aux_conciliation[
            "prev_discrepancies"
        ].items():
            if tipo in aux_discrepancy_rows:
                info = aux_discrepancy_rows[tipo]
                for _ in range(count):
                    col_index = info["index"] // 6
                    row_offset = info["index"] % 6

                    if col_index < len(info["col"]):
                        cell = (
                            f"{info['col'][col_index]}{info['row_start'] + row_offset}"
                        )
                        hoja[cell] = monto
                        hoja[cell].fill = orangeHighlighter
                        info["index"] += 1

        nombre_archivo = "conciliacion.xlsx"
        wb.save(nombre_archivo)

        # Devolver el archivo para descarga
        return FileResponse(
            path=nombre_archivo,
            filename=nombre_archivo,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        tb_str = traceback.format_exc()

        return JSONResponse(
            content={"error": f"Error al crear conciliación: {tb_str}"}, status_code=500
        )
